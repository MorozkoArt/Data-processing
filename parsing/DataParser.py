from common.listsCompare import listsCompare
from common.allChildren import allChildren
from parsing.parser.Parser import Parser
from common.constants import PARSED_HEADINGS, PARSED_NUMBER_FORMAT, PARSED_SHEET_NAME
from openpyxl.styles import Alignment, Border, Side


class DataParser:
    def __init__(self, root, subfolders, param, workbook):
        parser = Parser(root, subfolders, param)
        self.param = param
        self.parsed = parser.parseParamData()
        self.bigData = parser.convertToBigData(self.parsed)
        self.sheet = workbook.active
        self.sheet.title = PARSED_SHEET_NAME
    def logOutput(self):
        self.logHeading(self.param, {"col": 0, "row": 0}, len(allChildren(PARSED_HEADINGS)), PARSED_HEADINGS)
    def logHeading(self, value, offset, width, children=None, address=[]):
        cell = self.sheet.cell(row=1+offset["row"], column=1+offset["col"], value=value)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center', vertical='center')
        self.sheet.merge_cells(start_row=1+offset["row"], start_column=1+offset["col"], end_row=1+offset["row"], end_column=offset["col"]+width)
        colOffset = 0
        if type(children).__name__ == "list":
            for subvalue in children:
                self.logHeading(subvalue, {"row": offset["row"] + 1, "col": offset["col"]+colOffset}, 1, address=address+[subvalue])
                colOffset += 1
        if type(children).__name__ == "dict":
            for subvalue, subchildren in children.items():
                self.logHeading(subvalue, {"row": offset["row"] + 1, "col": offset["col"]+colOffset}, len(allChildren(subchildren)), subchildren, address + [subvalue])
                colOffset += len(allChildren(subchildren))
        if children is None:
            for addr in self.parsed.keys():
                if(listsCompare(addr, address)):
                    self.logColumn(addr, {"row": offset["row"] + 1, "col": offset["col"]})
    def logColumn(self, address, offset):
        rowIndex = 0
        for value in self.parsed[address]:
            cell = self.sheet.cell(row=1+rowIndex+offset["row"], column=1+offset["col"], value=value)
            cell.number_format = PARSED_NUMBER_FORMAT
            rowIndex += 1

    def getBigData(self):
        return self.bigData