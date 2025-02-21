from openpyxl.styles import Alignment, PatternFill, Font
from common.constants import ANOMALIES_BETA, ANOMALIES_BLOCKS_OFFSETS, ANOMALIES_COMPARABLE_FUNC, ANOMALIES_DATA_HEADINGS, ANOMALIES_DATA_VALUES, ANOMALIES_ERROR_BACKGROUND_COLOR, ANOMALIES_ERROR_TEXT_COLOR, ANOMALIES_NORMAL_NUMBER_FORMAT, ANOMALIES_NUMBER_FORMAT_CONDITIONS, ANOMALIES_PARAMS_HEADINGS, ANOMALIES_PARAMS_VALUES, ANOMALIES_STRICT_NUMBER_FORMAT, ANOMALIES_SUCCESS_BACKGROUND_COLOR, ANOMALIES_SUCCESS_TEXT_COLOR

class ExcelLogger:
    def __init__(self, column, sheet, offset):
        self.column = column
        self.sheet = sheet
        self.offset = offset
    def logColumn(self, success=False):
        self.logParams(success)
        self.logData(success)
    def logParams(self, success=False):
        offset = ANOMALIES_BLOCKS_OFFSETS[type(self.column).__name__]["params"]
        for index, heading in enumerate(ANOMALIES_PARAMS_HEADINGS):
            cell = self.sheet.cell(row=1+self.offset["row"]+offset["row"],column=1+index+self.offset["col"]+offset["col"], value=heading)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for index, function in enumerate(ANOMALIES_PARAMS_VALUES):
            value = getattr(self.column, function)()
            cell = self.sheet.cell(row=2+self.offset["row"]+offset["row"],column=1+index+self.offset["col"]+offset["col"], value=value)
            if ANOMALIES_NUMBER_FORMAT_CONDITIONS["from"] <= abs(cell.value) <= ANOMALIES_NUMBER_FORMAT_CONDITIONS["to"]:
                cell.number_format = ANOMALIES_NORMAL_NUMBER_FORMAT
            else:
                cell.number_format = ANOMALIES_STRICT_NUMBER_FORMAT
            if success:
                cell.fill = PatternFill(start_color=ANOMALIES_SUCCESS_BACKGROUND_COLOR, fill_type = "solid")
                cell.font = Font(color = ANOMALIES_SUCCESS_TEXT_COLOR)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    def logData(self, success=False):
        offset = ANOMALIES_BLOCKS_OFFSETS[type(self.column).__name__]["data"]
        for index, heading in enumerate(ANOMALIES_DATA_HEADINGS[type(self.column).__name__]):
            cell = self.sheet.cell(row=1+self.offset["row"]+offset["row"],column=1+index+self.offset["col"]+offset["col"], value=heading)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for functionIndex, function in enumerate(ANOMALIES_DATA_VALUES[type(self.column).__name__]):
            list = getattr(self.column, function)()
            for valueIndex, value in enumerate(list):
                cell = self.sheet.cell(row=2+valueIndex+self.offset["row"]+offset["row"],column=1+functionIndex+self.offset["col"]+offset["col"], value=value)
                if ANOMALIES_NUMBER_FORMAT_CONDITIONS["from"] <= abs(cell.value) <= ANOMALIES_NUMBER_FORMAT_CONDITIONS["to"]:
                    cell.number_format = ANOMALIES_NORMAL_NUMBER_FORMAT
                else:
                    cell.number_format = ANOMALIES_STRICT_NUMBER_FORMAT
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if function == ANOMALIES_COMPARABLE_FUNC and value > ANOMALIES_BETA:
                    cell.fill = PatternFill(start_color=ANOMALIES_ERROR_BACKGROUND_COLOR, fill_type = "solid")
                    cell.font = Font(color = ANOMALIES_ERROR_TEXT_COLOR)