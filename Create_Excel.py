import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from anomalies.AnomaliesFilter import AnomaliesFilter

def Create_excel(Data, BigData, name, root_directory):
    workbook = Workbook()
#----------------Лист-1 "Данные" ----------------------------------|
    sheet = workbook.active
    sheet.title = "Данные"
    # -----------------Формируем шапку таблицы---------------------|
    """Присваимаем значения заголовкам"""
    sheet.cell(row=1, column=1, value=name)
    sheet.cell(row=2, column=1, value=4)
    sheet.cell(row=2, column=5, value=8)
    sheet.cell(row=2, column=8, value=9)
    sheet.cell(row=3, column=1, value="PMinDie")
    sheet.cell(row=3, column=3, value="PM")
    sheet.cell(row=3, column=5, value="PM")
    sheet.cell(row=3, column=8, value="PMinDie")
    sheet.cell(row=3, column=11, value="PM")
    sheet.cell(row=4, column=1, value=1)
    sheet.cell(row=4, column=2, value=2)
    sheet.cell(row=4, column=3, value=3)
    sheet.cell(row=4, column=4, value=4)
    sheet.cell(row=4, column=5, value=2)
    sheet.cell(row=4, column=6, value=3)
    sheet.cell(row=4, column=7, value=4)
    sheet.cell(row=4, column=8, value=1)
    sheet.cell(row=4, column=9, value=2)
    sheet.cell(row=4, column=10, value=3)
    sheet.cell(row=4, column=11, value=15)
    sheet.cell(row=4, column=12, value=16)
    sheet.cell(row=4, column=13, value=17)
    """Объединяем ячейки"""
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    sheet.merge_cells(start_row=2, start_column=5, end_row=2, end_column=7)
    sheet.merge_cells(start_row=2, start_column=8, end_row=2, end_column=13)
    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    sheet.merge_cells(start_row=3, start_column=3, end_row=3, end_column=4)
    sheet.merge_cells(start_row=3, start_column=5, end_row=3, end_column=7)
    sheet.merge_cells(start_row=3, start_column=8, end_row=3, end_column=10)
    sheet.merge_cells(start_row=3, start_column=11, end_row=3, end_column=13)
    """Стилизуем"""
    thick_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row_index in range(1, 5):
        for col_index in range(1, 14):
            cell = sheet.cell(row=row_index, column=col_index)
            cell.border = thick_border
            cell.alignment = Alignment(horizontal='center', vertical='bottom')

    bold_font = Font(bold=True)
    cell = sheet.cell(row=1, column=1)
    cell.font = bold_font
    # -------------------Заполянем данными основную часть таблицы-------------------------|

    for key in Data:
        number, folder_name, SYS_WAFERID = key
        if number == 4:
            if folder_name == "PM":
                if SYS_WAFERID == 3:
                    FillingInTheTable(3, Data[key], sheet)
                elif SYS_WAFERID == 4:
                    FillingInTheTable(4, Data[key], sheet)
            elif folder_name == "PMinDie":
                if SYS_WAFERID == 1:
                    FillingInTheTable(1, Data[key], sheet)
                elif SYS_WAFERID == 2:
                    FillingInTheTable(2, Data[key], sheet)
        elif number == 8:
            if SYS_WAFERID == 2:
                FillingInTheTable(5, Data[key], sheet)
            elif SYS_WAFERID == 3:
                FillingInTheTable(6, Data[key], sheet)
            elif SYS_WAFERID == 4:
                FillingInTheTable(7, Data[key], sheet)
        elif number == 9:
            if folder_name == "PM":
                if SYS_WAFERID == 15:
                    FillingInTheTable(11, Data[key], sheet)
                elif SYS_WAFERID == 16:
                    FillingInTheTable(12, Data[key], sheet)
                elif SYS_WAFERID == 17:
                    FillingInTheTable(13, Data[key], sheet)
            elif folder_name == "PMinDie":
                if SYS_WAFERID == 1:
                    FillingInTheTable(8, Data[key], sheet)
                elif SYS_WAFERID == 2:
                    FillingInTheTable(9, Data[key], sheet)
                elif SYS_WAFERID == 3:
                    FillingInTheTable(10, Data[key], sheet)
    #----------------Лист-2 "Аномальные" ----------------------------------|
    anomaliesFilter = AnomaliesFilter(BigData, workbook)
    anomaliesFilter.logOutput()



    Save_file(workbook, name, root_directory)

def FillingInTheTable(col_index, list_data, sheet):
    number_format = "0.00E+00"
    for row_index in range(5, len(list_data)+5):
        cell = sheet.cell(row=row_index, column=col_index, value=list_data[row_index-5])
        cell.number_format = number_format
        cell.alignment = Alignment(vertical='bottom')

def Save_file(workbook, name, root_directory):
    filename = "ТК-ТХ503___Расчет___"+name+".xlsx"
    new_directory = "Results"
    full_path = os.path.join(root_directory, new_directory)
    if not os.path.exists(full_path):
        os.makedirs(full_path)
    filepath = os.path.join(full_path, filename)
    workbook.save(filepath)
    print(f"Файл успешно сохранен. ({filepath})")


